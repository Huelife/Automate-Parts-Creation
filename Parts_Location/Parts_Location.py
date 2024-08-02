#Install py -m pip install pyautogui
#Install py -m pip install openpyxl
#Install py -m pip install pyperclip
#Approx. 10 seconds/part

import pyautogui
import time
import pyperclip

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx", read_only = True)
sheet = workbook["UPLOAD 005"]
cell_value = ""

#Beginning actions
target_main = pyautogui.locateCenterOnScreen("Capture_Main.png",confidence=0.85)
pyautogui.click(target_main)     #Move the mouse to target_main coordinates and click

#Creating new parts
try:
    for X in range(1, 6):           #For loop, first value is 1, last value is 5 (6 is not counted)
        #New
        time.sleep(1)          #Sleep for 1 second
        target_1 = pyautogui.locateCenterOnScreen("Capture_01.png",confidence=0.85)
        pyautogui.click(target_1)     #Move the mouse to target_1 coordinates and click

        #Part ID
        cell_value = sheet.cell(row = X+2, column = 1).value
        pyperclip.copy(cell_value)          #Get Part ID
        time.sleep(1)          #Sleep for 1 second
        pyautogui.write(pyperclip.paste())          #Paste Part ID

        #Inventory Location ID
        cell_value = sheet.cell(row = X+2, column = 7).value
        pyperclip.copy(cell_value)          #Get Inventory Location ID
        time.sleep(0.2)          #Sleep for 0.2 second
        target_3 = pyautogui.locateCenterOnScreen("Capture_03.png",confidence=0.85)
        pyautogui.click(target_3)     #Move the mouse to target_3 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Inventory Location ID

        #Unit of Measure
        cell_value = sheet.cell(row = X+2, column = 13).value
        pyperclip.copy(cell_value)          #Get Unit of Measure
        time.sleep(0.2)          #Sleep for 0.2 second
        target_4 = pyautogui.locateCenterOnScreen("Capture_04.png",confidence=0.85)
        pyautogui.click(target_4)     #Move the mouse to target_4 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Unit of Measure

        #Bin ID
        cell_value = sheet.cell(row = X+2, column = 8).value
        pyperclip.copy(cell_value)          #Get Bin ID
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_5 = pyautogui.locateCenterOnScreen("Capture_05.png",confidence=0.85)
        pyautogui.click(target_5)     #Move the mouse to target_5 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Bin ID

        #Primary
        time.sleep(1)          #Sleep for 1 second
        target_6 = pyautogui.locateCenterOnScreen("Capture_06.png",confidence=0.85)
        pyautogui.click(target_6)     #Move the mouse to target_6 coordinates and click

        #Purchasing Info
        time.sleep(0.2)          #Sleep for 0.2 second
        target_7 = pyautogui.locateCenterOnScreen("Capture_07.png",confidence=0.85)
        pyautogui.click(target_7)     #Move the mouse to target_7 coordinates and click

        #Manufacturer Name
        cell_value = sheet.cell(row = X+2, column = 9).value
        pyperclip.copy(cell_value)          #Get Manufacturer Name
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_8 = pyautogui.locateCenterOnScreen("Capture_08.png",confidence=0.85)
        if cell_value == None:
            cell_value = "None"
        else:
            pyautogui.click(target_8)     #Move the mouse to target_8 coordinates and click       
            pyautogui.write(pyperclip.paste())          #Paste Manufacturer Name

        #Preferred Vendor ID
        cell_value = sheet.cell(row = X+2, column = 11).value
        pyperclip.copy(cell_value)          #Get Preferred Vendor ID
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_9 = pyautogui.locateCenterOnScreen("Capture_09.png",confidence=0.85)
        if cell_value == None:
            cell_value = "None"
        else:
            pyautogui.click(target_9)     #Move the mouse to target_9 coordinates and click       
            pyautogui.write(pyperclip.paste())          #Paste Preferred Vendor ID

        #Populate entry
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_10 = pyautogui.locateCenterOnScreen("Capture_10.png",confidence=0.85)
        pyautogui.click(target_10)     #Move the mouse to target_10 coordinates and click       

        #Cancel/Save
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_17 = pyautogui.locateCenterOnScreen("Capture_17.png",confidence=0.85)
        pyautogui.click(target_17)     #Move the mouse to target_17 coordinates and click 
        pyautogui.press("enter")     #Press the enter key
        time.sleep(2)          #Sleep for 2 seconds
except:
    print("ERROR")

workbook.close()
