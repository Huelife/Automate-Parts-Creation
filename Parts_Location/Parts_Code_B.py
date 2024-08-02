#Install py -m pip install openpyxl

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx", read_only = True)
sheet = workbook["UPLOAD 005"]
cell_value = ""

#For loop, first value is 1, last value is 3 (4 is not counted)
for X in range(1, 4):
    #Part ID
    cell_value = sheet.cell(row = X+2, column = 1).value
    print(cell_value)

    #Inventory Location ID
    cell_value = sheet.cell(row = X+2, column = 7).value
    print(cell_value)

    #Unit of Measure
    cell_value = sheet.cell(row = X+2, column = 13).value
    print(cell_value)                

    #Bin ID
    cell_value = sheet.cell(row = X+2, column = 8).value
    print(cell_value)

    #Manufacturer Name
    cell_value = sheet.cell(row = X+2, column = 9).value
    print(cell_value)

    #Preferred Vendor ID
    cell_value = sheet.cell(row = X+2, column = 11).value
    print(cell_value)

workbook.close()
