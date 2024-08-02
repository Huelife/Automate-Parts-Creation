#Install py -m pip install pyautogui
#Install py -m pip install openpyxl
#Install py -m pip install pyperclip
#Approx. 10 seconds/part

import pyautogui
import time
import pyperclip

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx", read_only = True)
sheet = workbook["UPLOAD 005"]
cell_value = ""

#Beginning actions
target_main = pyautogui.locateCenterOnScreen("Capture_Main.png",confidence=0.85)
pyautogui.click(target_main)     #Move the mouse to target_main coordinates and click

#Creating new parts
try:
    for X in range(1, 6):           #For loop, first value is 1, last value is 5 (6 is not counted)
        #New
        time.sleep(1)          #Sleep for 1 second
        target_1 = pyautogui.locateCenterOnScreen("Capture_01.png",confidence=0.85)
        pyautogui.click(target_1)     #Move the mouse to target_1 coordinates and click

        #Part ID
        cell_value = sheet.cell(row = X+2, column = 1).value
        pyperclip.copy(cell_value)          #Get Part ID
        time.sleep(2)          #Sleep for 2 second
        target_2 = pyautogui.locateCenterOnScreen("Capture_02.png",confidence=0.85)
        pyautogui.click(target_2)     #Move the mouse to target_2 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Part ID

        #Part Suffix
        cell_value = sheet.cell(row = X+2, column = 2).value
        pyperclip.copy(cell_value)          #Get Part Suffix
        time.sleep(0.2)          #Sleep for 0.2 second
        target_3 = pyautogui.locateCenterOnScreen("Capture_03.png",confidence=0.85)
        pyautogui.click(target_3)     #Move the mouse to target_3 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Part Suffix

        #Clear popup
        target_0 = pyautogui.locateCenterOnScreen("Capture_00.png",confidence=0.85)
        pyautogui.click(target_0)     #Move the mouse to target_0 coordinates and click

        #Keyword
        cell_value = sheet.cell(row = X+2, column = 3).value
        pyperclip.copy(cell_value)          #Get Keyword
        time.sleep(0.2)          #Sleep for 0.2 second
        target_4 = pyautogui.locateCenterOnScreen("Capture_04.png",confidence=0.85)
        pyautogui.click(target_4)     #Move the mouse to target_4 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Keyword

        #Clear popup
        time.sleep(0.2)          #Sleep for 0.2 seconds
        pyautogui.click(target_0)     #Move the mouse to target_0 coordinates and click

        #Short Description
        cell_value = sheet.cell(row = X+2, column = 4).value
        pyperclip.copy(cell_value)          #Get Short Description
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_5 = pyautogui.locateCenterOnScreen("Capture_05.png",confidence=0.85)
        pyautogui.click(target_5)     #Move the mouse to target_5 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Short Description

        #Product Category ID
        cell_value = sheet.cell(row = X+2, column = 5).value
        pyperclip.copy(cell_value)          #Get Product Category ID
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_6 = pyautogui.locateCenterOnScreen("Capture_06.png",confidence=0.85)
        pyautogui.click(target_6)     #Move the mouse to target_6 coordinates and click       
        pyautogui.write(pyperclip.paste())          #Paste Product Category ID

        #Part Classification ID
        cell_value = sheet.cell(row = X+2, column = 6).value
        pyperclip.copy(cell_value)          #Get Part Classification ID
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_7 = pyautogui.locateCenterOnScreen("Capture_07.png",confidence=0.85)
        pyautogui.click(target_7)     #Move the mouse to target_7 coordinates and click       
        pyautogui.write(pyperclip.paste())          #Paste Part Classification ID

        #Cancel/Save
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_8 = pyautogui.locateCenterOnScreen("Capture_08A.png",confidence=0.85)
        pyautogui.click(target_8)     #Move the mouse to target_8 coordinates and click 
        pyautogui.press("enter")     #Press the enter key
        time.sleep(2)          #Sleep for 2 seconds
except:
    print("ERROR")

workbook.close()
