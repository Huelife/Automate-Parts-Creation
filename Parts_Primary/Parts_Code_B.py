#Install py -m pip install openpyxl

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx")
sheet = workbook["UPLOAD 005"]
cell_value = ""

#For loop, first value is 1, last value is 3 (4 is not counted)
for X in range(1, 4):
    #Part ID
    cell_value = sheet.cell(row = X+2, column = 1).value
    print(cell_value)

    #Part Suffix
    cell_value = sheet.cell(row = X+2, column = 2).value
    print(cell_value)

    #Keyword
    cell_value = sheet.cell(row = X+2, column = 3).value
    print(cell_value)

    #Short Description
    cell_value = sheet.cell(row = X+2, column = 4).value
    print(cell_value)                

    #Product Category ID
    cell_value = sheet.cell(row = X+2, column = 5).value
    print(cell_value)
    
    #Part Classification ID
    cell_value = sheet.cell(row = X+2, column = 6).value
    print(cell_value)
