#Install py -m pip install openpyxl

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx", read_only = True)
sheet = workbook["UPLOAD 005"]
cell_value = ""

#For loop, first value is 1, last value is 3 (4 is not counted)
for X in range(1, 4):
    #Inventory Location ID
    cell_value = sheet.cell(row = X+2, column = 7).value
    print(cell_value)
    
    #Part ID
    cell_value = sheet.cell(row = X+2, column = 1).value
    print(cell_value)

    #Inventory Count
    cell_value = sheet.cell(row = X+2, column = 12).value
    print(cell_value)

    #Price
    cell_value = sheet.cell(row = X+2, column = 14).value
    print(cell_value)

workbook.close()
