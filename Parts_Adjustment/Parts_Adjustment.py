#Install py -m pip install pyautogui
#Install py -m pip install openpyxl
#Install py -m pip install pyperclip
#Approx. 10 seconds/part

import pyautogui
import time
import pyperclip

from openpyxl import load_workbook

#Declare variables
workbook = load_workbook(filename = "Parts Inventory Master Sheet.xlsx", read_only = True)
sheet = workbook["UPLOAD 005"]
cell_value = ""

#Beginning actions
target_main = pyautogui.locateCenterOnScreen("Capture_Main.png",confidence=0.85)
pyautogui.click(target_main)     #Move the mouse to target_main coordinates and click

#Creating new parts
try:
    for X in range(1, 6):           #For loop, first value is 1, last value is 5 (6 is not counted)
        #New
        time.sleep(1)          #Sleep for 1 second
        target_1 = pyautogui.locateCenterOnScreen("Capture_01.png",confidence=0.85)
        pyautogui.click(target_1)     #Move the mouse to target_1 coordinates and click

        #Inventory Location ID
        cell_value = sheet.cell(row = X+2, column = 7).value
        pyperclip.copy(cell_value)          #Get Inventory Location ID
        time.sleep(1)          #Sleep for 1 second
        pyautogui.write(pyperclip.paste())          #Paste Inventory Location ID
        time.sleep(0.2)          #Sleep for 0.2 second

        #Part ID
        cell_value = sheet.cell(row = X+2, column = 1).value
        pyperclip.copy(cell_value)          #Get Part ID
        time.sleep(1)          #Sleep for 1 second
        target_2 = pyautogui.locateCenterOnScreen("Capture_02.png",confidence=0.85)
        pyautogui.click(target_2)     #Move the mouse to target_2 coordinates and click
        pyautogui.write(pyperclip.paste())          #Paste Part ID

        #Populate entry
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_00 = pyautogui.locateCenterOnScreen("Capture_00.png",confidence=0.85)
        pyautogui.click(target_00)     #Move the mouse to target_00 coordinates and click

        #Change
        time.sleep(0.2)          #Sleep for 0.2 second
        target_3 = pyautogui.locateCenterOnScreen("Capture_03.png",confidence=0.85)
        pyautogui.click(target_3)     #Move the mouse to target_3 coordinates and click

        #If statement for Inventory Count
        cell_value = sheet.cell(row = X+2, column = 12).value
        pyperclip.copy(cell_value)          #Get Inventory Count
        time.sleep(0.2)          #Sleep for 0.2 seconds
        if cell_value == None:
            #Unit Price Only
            time.sleep(0.2)          #Sleep for 0.2 seconds
            target_4B = pyautogui.locateCenterOnScreen("Capture_04B.png",confidence=0.85)
            pyautogui.click(target_4B)     #Move the mouse to target_4B coordinates and click

            #Price
            cell_value = sheet.cell(row = X+2, column = 14).value
            pyperclip.copy(cell_value)
            time.sleep(0.2)          #Sleep for 0.2 seconds
            target_6 = pyautogui.locateCenterOnScreen("Capture_06.png",confidence=0.85)
            pyautogui.click(target_6)     #Move the mouse to target_6 coordinates and click
            pyautogui.write(pyperclip.paste())          #Paste Price
        else:
            #Qty at a Different Price
            time.sleep(0.2)          #Sleep for 0.2 seconds
            target_4A = pyautogui.locateCenterOnScreen("Capture_04A.png",confidence=0.85)
            pyautogui.click(target_4A)     #Move the mouse to target_4A coordinates and click

            #Inventory Count
            cell_value = sheet.cell(row = X+2, column = 12).value
            pyperclip.copy(cell_value)          #Get Inventory Count
            time.sleep(1)          #Sleep for 1 seconds
            target_5 = pyautogui.locateCenterOnScreen("Capture_05.png",confidence=0.85)
            pyautogui.click(target_5)     #Move the mouse to target_5 coordinates and click
            pyautogui.write(pyperclip.paste())          #Paste Inventory Count

            #Price
            cell_value = sheet.cell(row = X+2, column = 14).value
            pyperclip.copy(cell_value)          #Get Price
            time.sleep(0.2)          #Sleep for 0.2 seconds
            target_6 = pyautogui.locateCenterOnScreen("Capture_06.png",confidence=0.85)
            pyautogui.click(target_6)     #Move the mouse to target_6 coordinates and click
            time.sleep(0.2)          #Sleep for 0.2 seconds
            if cell_value == None:
                pyperclip.copy("0.00")          #Get Price when value is None
                pyautogui.write(pyperclip.paste())          #Paste Price
            else:
                pyautogui.write(pyperclip.paste())          #Paste Price

        #Reason Code
        pyperclip.copy("CYCLE CT")
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_7 = pyautogui.locateCenterOnScreen("Capture_07.png",confidence=0.85)
        pyautogui.click(target_7)     #Move the mouse to target_7 coordinates and click       
        pyautogui.write(pyperclip.paste())          #Paste Reason Code

        #Comments
        pyperclip.copy("BATCH UPLOAD MANUAL ADJUSTMENT")
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_8 = pyautogui.locateCenterOnScreen("Capture_08.png",confidence=0.85)
        pyautogui.click(target_8)     #Move the mouse to target_8 coordinates and click       
        pyautogui.write(pyperclip.paste())          #Paste Comments

        #Cancel/Save
        time.sleep(0.2)          #Sleep for 0.2 seconds
        target_10 = pyautogui.locateCenterOnScreen("Capture_10.png",confidence=0.85)
        pyautogui.click(target_10)     #Move the mouse to target_10 coordinates and click 
        pyautogui.press("enter")     #Press the enter key
        time.sleep(2)          #Sleep for 2 seconds
except:
    print("ERROR")

workbook.close()
